import tkinter as tk
import tkinter.ttk as ttk
from tkinter import *
from tkinter.ttk import *
from turtle import color, onclick
import random
import time
import os
import pandas as pd
import openpyxl as xl
from openpyxl import Workbook

# Define all the functions

def find_center(x, y, w, h):
    return x + w//2, y + h//2

def find_center_rect(x1, x2, y1, y2):
    return x1 + (x2 - x1) // 2, y1 + (y2 - y1) // 2 + 50

def check_correctness(rl, ec):
    if ec=="1":
        if rl in ["A", "B", "C", "D", "E", "F"]:
            correctness = True
        else:
            correctness = False
    elif ec=="2":
        if rl in ["G", "H", "I", "J", "K", "L"]:
            correctness = True
        else:
            correctness = False
    elif ec=="3":
        if rl in ["M", "N", "O", "P", "Q", "R"]:
            correctness = True
        else:
            correctness = False
    elif ec=="4":
        if rl in ["S", "T", "U", "V", "W", "X", "Y", "Z"]:
            correctness = True
        else:
            correctness = False
    else:
        correctness = False
    return correctness

def make_button_inv():
    global i, txt, tw, txt2, tw2, start_time, end_time, elapsed_time, test_button, button_visible
    if button_visible == True:
        test_button.place_forget()
        button_visible = False
        txt="Find the following letter:"
        c.itemconfig(tw, text=txt)
        txt2 = random_letters[i]
        # Start the stopwatch
        start_time = time.time()
        if i==0:
            tw2 = c.create_text(700-25, 70, text=txt2, font=("Courier", 50), fill="white")
        else:
            c.itemconfig(tw2, text=txt2)
    # Call the key press listener
    master.bind("<KeyPress>", on_key_press)

def on_key_press(event):
    global i, txt, tw, txt2, tw2, start_time, end_time, elapsed_time, test_button, button_visible, result
    if button_visible == False:
        # Stop the stopwatch
        end_time = time.time()
        # Calculate elapsed time
        elapsed_time = end_time - start_time
        if i<12:
            correctness = check_correctness(random_letters[i], event.char)
            print(f"i: {i} - letter: {random_letters[i]} - key pressed: {event.char} - elapsed time: {elapsed_time} - correctness: {correctness}")
            result.append(i)
            result.append(random_letters[i])
            result.append(event.char)
            result.append(elapsed_time)
            result.append(correctness)

            # Save the experiment data to Excel
            # Convert the list to a DataFrame
            df = pd.DataFrame([result])
            # Define the Excel file name
            file_name = "data.xlsx"

            # Check if the file not exists
            if os.path.isfile(file_name)==False:
                # Create a new Excel file
                wb = Workbook()
                ws = wb.active
                ws.title = "Sheet1"
                # Save the new Excel file
                wb.save(filename=file_name)

            # Get number of rows in excel file (to determine where to append)
            source_file = xl.load_workbook("data.xlsx", enumerate)
            sheet = source_file["Sheet1"]
            row_count = sheet.max_row
            source_file.close()
            with pd.ExcelWriter("data.xlsx", mode='a', if_sheet_exists='overlay') as writer:  
                df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow = row_count)

            result = []
            test_button.place(x=650, y=70)

            if i==11: # finish
                print("Finish")
                c.delete(tw)
                c.delete(tw2)
                tw2 = c.create_text(700-25, 50, text="End of the experiment - Thank you!", font=("Courier", 30) ,fill="white")
                test_button.place_forget()
                button_visible = False
            else:
                txt="Press start to begin"
                c.itemconfig(tw, text=txt)
                c.itemconfig(tw2, text="")
                button_visible = True
        i = i + 1

# Choose 12 random letters for the Tscan experiment
Q1 = ['A', 'B', 'C', 'D', 'E', 'F']
Q2 = ['G', 'H', 'I', 'J', 'K', 'L']
Q3 = ['M', 'N', 'O', 'P', 'Q', 'R']
Q4 = ['S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
random_letters_Q1 = random.sample(Q1, 3)
random_letters_Q2 = random.sample(Q2, 3)
random_letters_Q3 = random.sample(Q3, 3)
random_letters_Q4 = random.sample(Q4, 3)
# random_letters = [item for sublist in [random_letters_Q1, random_letters_Q2, random_letters_Q3, random_letters_Q4] for item in sublist]
# random.shuffle(random_letters)
# print(random_letters)
# Modified - The first 4 letters will be taken from different quarters, the rest order will be free
random_letters_group1 = [random_letters_Q1[0],random_letters_Q2[0],random_letters_Q3[0],random_letters_Q4[0]]
random.shuffle(random_letters_group1)
random_letters_group2 = [random_letters_Q1[1],random_letters_Q2[1],random_letters_Q3[1],random_letters_Q4[1],random_letters_Q1[2],random_letters_Q2[2],random_letters_Q3[2],random_letters_Q4[2]]
random.shuffle(random_letters_group2)
random_letters = random_letters_group1 + random_letters_group2
print(random_letters)

# GUI
master=tk.Tk()
master.title("Tscan Experiment - Gaze Controlled Keyboard inspired by SSVEP Methods")

height = 900
width = 1400
x = (master.winfo_screenwidth()//2)-(width//2) 
y = (master.winfo_screenheight()//2)-(height//2) 
master.geometry("{}x{}+{}+{}".format(width, height, x, y)) 

txt = tk.Label(master, text="", bg="black", fg="white", font=("Times New Roman", 20))
overall_str = ""

c = tk.Canvas(master, width=1400, height=800, bg="black")
c.pack(pady=5, padx=10)

# Setting for button
button_visible = True
test_button = ttk.Button(c, text="START", command=make_button_inv)
# test_button.place(x=700 - 175 - 500, y=20)
test_button.place(x=650, y=70)

### Q1 ###
w  = 700 - 175 - 500
h = 75 + 25
r1 = c.create_rectangle(w, h, w + 500, h + 275, fill="#b5b4b1")
x, y = find_center(w, h, 500, 275)
t1 = c.create_text(x, y, text="A, B, C, D, E, F", font=("Courier", 30), fill="black")
t1_cmd = c.create_text(w + 70, h + 20, text="Press '1'", font=("Courier", 20), fill="red")

### Q2 ###
w  = 700 + 150
h = 75 + 25
r3 = c.create_rectangle(w, h, w + 500, h + 275, fill="#b5b4b1")
x, y = find_center(w, h, 500, 275)
t3 = c.create_text(x, y, text="G, H, I, J, K, L", font=("Courier", 30), fill="black")
t3_cmd = c.create_text(w + 70, h + 20, text="Press '2'", font=("Courier", 20), fill="red")

### Q3 ###
w  = 700 - 175 - 500
h = 800 - 25 - 300 - 15
r2 = c.create_rectangle(w, h, w + 500, h + 275, fill="#b5b4b1")
x, y = find_center(w, h, 500, 275)
t2 = c.create_text(x, y, text="M, N, O, P, Q, R", font=("Courier", 30), fill="black")
t2_cmd = c.create_text(w + 70, h + 20, text="Press '3'", font=("Courier", 20), fill="red")

### Q4 ###
w  = 700 + 150 
h = 800 - 25 - 300 - 15
r4 = c.create_rectangle(w, h, w + 500, h + 275, fill="#b5b4b1")
x, y = find_center(w, h, 500, 275)
t4 = c.create_text(x, y, text="S, T, U, V, W, X, Y, Z", font=("Courier", 30), fill="black")
t4_cmd = c.create_text(w + 70, h + 20, text="Press '4'", font=("Courier", 20), fill="red")

### Neutral Zone ###
x, y = find_center_rect(700-175, 700+150, 275,  800 - 25 - 300 - 15)
tnz = c.create_text(x, y, text="Neutral Zone", font=("Courier", 30), fill="white")

### Instructions ###
txt="Press start to begin the experiment"
tw = c.create_text(700-25, 20, text=txt, font=("Courier", 30), fill="white")    

# Iteration 1
i = 0
result = []

master.configure(background='black')
master.mainloop()